import io
import qrcode
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import make_response
from datetime import datetime, date
from models import *
import pytz

def generate_qr_code(data):
    """Generate QR code for given data"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    # Convert to base64 for display
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_str = base64.b64encode(img_buffer.getvalue()).decode()

    return f"data:image/png;base64,{img_str}"

def calculate_sop_ingredients(sop_id, kg_quantity):
    """Calculate ingredient quantities based on SOP and requested kg"""
    from models import SOP, SOPIngredient, Ingredient
    from app import db

    sop = SOP.query.get(sop_id)
    if not sop:
        print(f"SOP with ID {sop_id} not found")
        return []

    # Get SOP ingredients using proper join
    sop_ingredients = db.session.query(SOPIngredient).filter_by(sop_id=sop_id).all()

    if not sop_ingredients:
        print(f"No ingredients found for SOP {sop_id}")
        return []

    # Calculate scaling factor
    scale_factor = kg_quantity / sop.base_quantity
    print(f"Scale factor: {scale_factor} (target: {kg_quantity} kg, base: {sop.base_quantity} kg)")

    ingredients = []
    for sop_ingredient in sop_ingredients:
        ingredient = Ingredient.query.get(sop_ingredient.ingredient_id)
        if ingredient:
            ingredients.append({
                'ingredient': ingredient,
                'quantity': sop_ingredient.quantity * scale_factor,
                'note': sop_ingredient.note
            })

    print(f"Calculated {len(ingredients)} ingredients for SOP {sop_id}")
    return ingredients

def update_inventory_after_cooking(sop_id, kg_quantity, kitchen_id):
    """Automatically update inventory after cooking based on SOP"""
    from datetime import date
    from models import InventoryTracker, Ingredient
    from app import db

    ingredients = calculate_sop_ingredients(sop_id, kg_quantity)
    today = date.today()

    for ingredient_data in ingredients:
        ingredient = ingredient_data['ingredient']
        used_qty = ingredient_data['quantity']

        # Auto-create tracker if doesn't exist
        tracker = auto_create_inventory_tracker(kitchen_id, ingredient.id, today)

        if tracker:
            # Add usage to tracker
            tracker.used_qty += used_qty
            tracker.closing_stock = tracker.opening_stock + tracker.purchase_qty - tracker.used_qty - tracker.loss_qty

            # Prevent negative stock
            if tracker.closing_stock < 0:
                tracker.closing_stock = 0

            # Update status
            if tracker.closing_stock <= 0:
                tracker.status = 'Critical'
            elif tracker.closing_stock <= tracker.min_threshold:
                tracker.status = 'Low'
            else:
                tracker.status = 'OK'

    db.session.commit()

def export_table_to_excel(table_name):
    """Export database table to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = table_name

    # Define table mappings
    table_models = {
        'users': User,
        'kitchens': Kitchen,
        'ingredients': Ingredient,
        'sops': SOP,
        'refill_requests': RefillRequest,
        'daily_sales': DailySales,
        'inventory_tracker': InventoryTracker,
        'supplier_purchases': SupplierPurchase,
        'customer_feedback': CustomerFeedback,
        'inventory_adjustments': InventoryAdjustment
    }

    if table_name not in table_models:
        return None

    model = table_models[table_name]

    # Get column names
    columns = [column.name for column in model.__table__.columns]

    # Write headers
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data
    data = model.query.all()
    for row_num, record in enumerate(data, 2):
        for col_num, column in enumerate(columns, 1):
            value = getattr(record, column)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, date):
                value = value.strftime('%Y-%m-%d')
            ws.cell(row=row_num, column=col_num, value=value)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={table_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'

    return response

def calculate_daily_sales_totals(form_data, cart_staff, sop):
    """Calculate all totals for daily sales entry"""
    kg_unsold = float(form_data['kg_unsold']) if form_data['kg_unsold'] is not None else 0.0
    cash_collected = float(form_data['cash_collected']) if form_data['cash_collected'] is not None else 0.0

    # Get today's total refill requests for this cart staff and recipe
    today = date.today()
    total_kg_taken = db.session.query(db.func.sum(RefillRequest.kg_request)).filter_by(
        cart_staff_id=cart_staff.id,
        sop_id=sop.id,
        status='delivered'
    ).filter(db.func.date(RefillRequest.pickup_time) == today).scalar() or 0

    kg_sold = max(0, total_kg_taken - kg_unsold)
    total_revenue_expected = kg_sold * sop.recipe_rate
    upi_collected = max(0, total_revenue_expected - cash_collected)
    total_revenue = cash_collected + upi_collected
    total_incentive = kg_sold * cart_staff.incentive

    return {
        'kg_taken': total_kg_taken,
        'kg_sold': kg_sold,
        'upi_collected': upi_collected,
        'total_revenue': total_revenue,
        'recipe_rate': sop.recipe_rate,
        'incentive_per_kg': cart_staff.incentive,
        'total_incentive': total_incentive
    }

def get_kitchen_performance_data(kitchen_id, days=7):
    """Get performance data for a kitchen"""
    from datetime import timedelta

    end_date = date.today()
    start_date = end_date - timedelta(days=days-1)

    # Sales data
    sales_data = db.session.query(
        DailySales.date,
        db.func.sum(DailySales.total_revenue).label('revenue'),
        db.func.sum(DailySales.kg_sold).label('kg_sold')
    ).filter(
        DailySales.kitchen_id == kitchen_id,
        DailySales.date >= start_date,
        DailySales.date <= end_date
    ).group_by(DailySales.date).all()

    # Refill requests data
    refill_data = db.session.query(
        db.func.date(RefillRequest.request_time).label('date'),
        db.func.count(RefillRequest.id).label('requests'),
        db.func.avg(
            db.func.extract('epoch', RefillRequest.delivered_time - RefillRequest.pickup_time) / 60
        ).label('avg_fulfillment_time')
    ).filter(
        RefillRequest.kitchen_id == kitchen_id,
        RefillRequest.request_time >= start_date,
        RefillRequest.status == 'delivered'
    ).group_by(db.func.date(RefillRequest.request_time)).all()

    return {
        'sales_data': sales_data,
        'refill_data': refill_data
    }

def get_cart_performance_data(cart_staff_id, days=7):
    """Get performance data for a cart"""
    from datetime import timedelta

    end_date = date.today()
    start_date = end_date - timedelta(days=days-1)

    performance_data = db.session.query(
        DailySales.date,
        db.func.sum(DailySales.total_revenue).label('revenue'),
        db.func.sum(DailySales.kg_sold).label('kg_sold'),
        db.func.sum(DailySales.kg_unsold).label('kg_unsold'),
        db.func.sum(DailySales.total_incentive).label('incentive')
    ).filter(
        DailySales.cart_staff_id == cart_staff_id,
        DailySales.date >= start_date,
        DailySales.date <= end_date
    ).group_by(DailySales.date).all()

    return performance_data

def auto_create_inventory_tracker(kitchen_id, ingredient_id, date_obj=None):
    """Auto-create inventory tracker when needed"""
    from datetime import date

    if not date_obj:
        date_obj = date.today()

    # Check if tracker already exists
    existing = InventoryTracker.query.filter_by(
        date=date_obj,
        kitchen_id=kitchen_id,
        ingredient_id=ingredient_id
    ).first()

    if existing:
        return existing

    # Get ingredient for minimum stock
    ingredient = Ingredient.query.get(ingredient_id)
    if not ingredient:
        return None

    # Get previous day's closing stock
    yesterday = date.fromordinal(date_obj.toordinal() - 1)
    yesterday_tracker = InventoryTracker.query.filter_by(
        date=yesterday,
        kitchen_id=kitchen_id,
        ingredient_id=ingredient_id
    ).first()

    opening_stock = yesterday_tracker.closing_stock if yesterday_tracker else 0.0

    # Create new tracker
    tracker = InventoryTracker(
        date=date_obj,
        kitchen_id=kitchen_id,
        ingredient_id=ingredient_id,
        opening_stock=opening_stock,
        purchase_qty=0.0,
        used_qty=0.0,
        loss_qty=0.0,
        closing_stock=opening_stock,
        min_threshold=ingredient.minimum_stock,
        status='Critical' if opening_stock <= 0 else 'Low' if opening_stock <= ingredient.minimum_stock else 'OK'
    )
    db.session.add(tracker)
    return tracker

def initialize_daily_inventory_for_kitchen(kitchen_id, date_obj=None):
    """Initialize daily inventory for all ingredients in a kitchen"""
    from datetime import date
    from models import Ingredient, InventoryTracker
    from app import db
    
    if not date_obj:
        date_obj = date.today()
    
    ingredients = Ingredient.query.all()
    for ingredient in ingredients:
        auto_create_inventory_tracker(kitchen_id, ingredient.id, date_obj)
    
    db.session.commit()

def create_demo_data():
    """Create initial admin user only - inventory system is fully automatic"""
    if User.query.count() > 0:
        return

    # Create admin user (not tied to any specific kitchen)
    admin = User(
        username='admin',
        name='System Admin',
        role='admin',
        kitchen_id=None,  # Admin not tied to specific kitchen
        location='Head Office',
        mobile_no='9999999999',
        email='admin@company.com',
        address='Head Office Address',
        is_active=True
    )
    admin.set_password('admin123')  # Default password
    db.session.add(admin)

    db.session.commit()
    print("Demo data created - Automatic inventory system ready!")
    print("Admin user: admin / admin123")

def get_total_refills_completed(kitchen_id):
    """Get total refills completed today for a kitchen"""
    today = date.today()
    return RefillRequest.query.filter_by(
        kitchen_id=kitchen_id,
        status='delivered'
    ).filter(db.func.date(RefillRequest.pickup_time) == today).scalar() or 0

def get_average_delivery_time(kitchen_id):
    """Calculate average delivery time for a kitchen"""
    today = date.today()
    return db.session.query(
        db.func.avg(
            db.func.extract('epoch', RefillRequest.delivered_time - RefillRequest.pickup_time) / 60
        )
    ).filter_by(
        kitchen_id=kitchen_id,
        status='delivered'
    ).filter(
        RefillRequest.delivered_time.isnot(None),
        RefillRequest.pickup_time.isnot(None)
    ).scalar() or 0

def convert_utc_to_ist_in_db():
    """Convert all UTC timestamps in database to IST"""
    from app import db
    from models import RefillRequest, User, Kitchen, Ingredient, SOP, DailySales, SupplierPurchase, CustomerFeedback, InventoryAdjustment, Cart

    # Get timezone objects
    utc_tz = pytz.timezone('UTC')
    ist_tz = pytz.timezone('Asia/Kolkata')

    # Function to convert a single timestamp
    def convert_timestamp(timestamp):
        if timestamp is None:
            return None
        if timestamp.tzinfo is None:
            # Assume UTC if no timezone
            timestamp = utc_tz.localize(timestamp)
        ist_time = timestamp.astimezone(ist_tz)
        return ist_time.replace(tzinfo=None)

    # Convert timestamps in RefillRequest
    refill_requests = RefillRequest.query.all()
    for refill in refill_requests:
        refill.request_time = convert_timestamp(refill.request_time)
        refill.taken_by_chef_time = convert_timestamp(refill.taken_by_chef_time)
        refill.prepared_time = convert_timestamp(refill.prepared_time)
        refill.pickup_time = convert_timestamp(refill.pickup_time)
        refill.delivered_time = convert_timestamp(refill.delivered_time)

    # Convert created_at timestamps in all models
    for model in [User, Kitchen, Ingredient, SOP, DailySales, SupplierPurchase, CustomerFeedback, InventoryAdjustment, Cart]:
        items = model.query.all()
        for item in items:
            if hasattr(item, 'created_at'):
                item.created_at = convert_timestamp(item.created_at)

    # Commit all changes
    db.session.commit()

def get_indian_time():
    """Get current time in IST timezone"""
    indian_tz = pytz.timezone('Asia/Kolkata')
    return datetime.now(indian_tz).replace(tzinfo=None)